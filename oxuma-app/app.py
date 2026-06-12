import os, json, sqlite3
from datetime import datetime
from flask import Flask, render_template, jsonify, request, redirect, url_for, flash, send_file
import io

app = Flask(__name__)
app.secret_key = "oxuma2024secretkey"
app.jinja_env.filters['enumerate'] = enumerate
DB = os.path.join(os.path.dirname(__file__), "oxuma.db")

# ─── DATABASE ────────────────────────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    c = conn.cursor()
    c.executescript("""
        CREATE TABLE IF NOT EXISTS productos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            categoria TEXT,
            marca TEXT,
            precio_costo REAL DEFAULT 0,
            margen REAL DEFAULT 30,
            precio_venta REAL DEFAULT 0,
            stock INTEGER DEFAULT 0,
            activo INTEGER DEFAULT 1,
            updated_at TEXT
        );
        CREATE TABLE IF NOT EXISTS clientes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            cuit TEXT,
            email TEXT,
            telefono TEXT,
            direccion TEXT,
            tipo TEXT DEFAULT 'mayorista',
            created_at TEXT
        );
        CREATE TABLE IF NOT EXISTS pedidos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cliente_id INTEGER,
            fecha TEXT,
            total REAL DEFAULT 0,
            estado TEXT DEFAULT 'pendiente',
            notas TEXT,
            FOREIGN KEY (cliente_id) REFERENCES clientes(id)
        );
        CREATE TABLE IF NOT EXISTS pedido_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            pedido_id INTEGER,
            producto_id INTEGER,
            cantidad INTEGER DEFAULT 1,
            precio_unitario REAL,
            FOREIGN KEY (pedido_id) REFERENCES pedidos(id),
            FOREIGN KEY (producto_id) REFERENCES productos(id)
        );
        CREATE TABLE IF NOT EXISTS historial_precios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            producto_id INTEGER,
            precio_anterior REAL,
            precio_nuevo REAL,
            motivo TEXT,
            fecha TEXT,
            FOREIGN KEY (producto_id) REFERENCES productos(id)
        );
        CREATE TABLE IF NOT EXISTS precio_reglas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cantidad_minima INTEGER NOT NULL,
            descuento_pct REAL NOT NULL DEFAULT 0,
            etiqueta TEXT
        );
    """)
    # Reglas por defecto
    if conn.execute("SELECT COUNT(*) FROM precio_reglas").fetchone()[0] == 0:
        c.executemany(
            "INSERT INTO precio_reglas (cantidad_minima, descuento_pct, etiqueta) VALUES (?,?,?)",
            [(3, 5.0, "x3"), (12, 10.0, "x12"), (24, 15.0, "x24"), (48, 20.0, "x48")]
        )
    # Load products if table is empty
    count = c.execute("SELECT COUNT(*) FROM productos").fetchone()[0]
    if count == 0:
        json_path = os.path.join(os.path.dirname(__file__), "productos_raw.json")
        if os.path.exists(json_path):
            with open(json_path, "r", encoding="utf-8") as f:
                raw = json.load(f)
            productos = [p for p in raw if "nombre" in p and "precio" in p and p["precio"]]
            now = datetime.now().isoformat()
            MARGEN_DEFAULT = 30.0
            for p in productos:
                precio_venta = float(p["precio"])
                # precio_costo = precio_venta / (1 + margen/100)  → cada producto tiene precio individual
                precio_costo = round(precio_venta / (1 + MARGEN_DEFAULT / 100), 2)
                c.execute("""
                    INSERT INTO productos (nombre, categoria, marca, precio_costo, margen, precio_venta, stock, activo, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, 1, ?)
                """, (p["nombre"], p.get("categoria",""), p.get("marca",""), precio_costo, MARGEN_DEFAULT, precio_venta, 0, now))
    conn.commit()
    conn.close()

def migrate_db():
    """Agrega columnas nuevas y tablas sin destruir datos existentes."""
    conn = get_db()
    c = conn.cursor()
    # Nuevas columnas en productos
    nuevas_cols = [
        ("codigo",     "TEXT DEFAULT ''"),
        ("precio_x3",  "REAL DEFAULT 0"),
        ("precio_x12", "REAL DEFAULT 0"),
        ("precio_x24", "REAL DEFAULT 0"),
        ("precio_x48", "REAL DEFAULT 0"),
    ]
    for col, defn in nuevas_cols:
        try:
            c.execute(f"ALTER TABLE productos ADD COLUMN {col} {defn}")
        except Exception:
            pass
    # Tabla combos
    c.executescript("""
        CREATE TABLE IF NOT EXISTS combos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            descripcion TEXT,
            precio_combo REAL DEFAULT 0,
            activo INTEGER DEFAULT 1,
            ia_generado INTEGER DEFAULT 0,
            created_at TEXT
        );
        CREATE TABLE IF NOT EXISTS combo_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            combo_id INTEGER,
            producto_id INTEGER,
            cantidad INTEGER DEFAULT 1,
            FOREIGN KEY (combo_id) REFERENCES combos(id),
            FOREIGN KEY (producto_id) REFERENCES productos(id)
        );
        CREATE TABLE IF NOT EXISTS precio_reglas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cantidad_minima INTEGER NOT NULL,
            descuento_pct REAL NOT NULL DEFAULT 0,
            etiqueta TEXT
        );
    """)
    # Insertar reglas por defecto si la tabla está vacía
    existing = conn.execute("SELECT COUNT(*) as n FROM precio_reglas").fetchone()["n"]
    if existing == 0:
        conn.executemany(
            "INSERT INTO precio_reglas (cantidad_minima, descuento_pct, etiqueta) VALUES (?,?,?)",
            [(3, 5.0, "x3"), (12, 10.0, "x12"), (24, 15.0, "x24"), (48, 20.0, "x48")]
        )
    conn.commit()
    conn.close()

# ─── HELPERS ─────────────────────────────────────────────────────────────────

def query(sql, args=(), one=False):
    conn = get_db()
    cur = conn.execute(sql, args)
    rv = cur.fetchall()
    conn.close()
    return (rv[0] if rv else None) if one else rv

def execute(sql, args=()):
    conn = get_db()
    cur = conn.execute(sql, args)
    conn.commit()
    last_id = cur.lastrowid
    conn.close()
    return last_id

# ─── ROUTES ──────────────────────────────────────────────────────────────────

@app.route("/")
def dashboard():
    total_productos = query("SELECT COUNT(*) as n FROM productos WHERE activo=1", one=True)["n"]
    total_clientes = query("SELECT COUNT(*) as n FROM clientes", one=True)["n"]
    total_pedidos = query("SELECT COUNT(*) as n FROM pedidos", one=True)["n"]
    valor_catalogo = query("SELECT SUM(precio_venta) as s FROM productos WHERE activo=1", one=True)["s"] or 0

    # Products by category
    por_categoria = query("""
        SELECT categoria, COUNT(*) as cantidad, AVG(precio_venta) as precio_prom
        FROM productos WHERE activo=1
        GROUP BY categoria ORDER BY cantidad DESC LIMIT 12
    """)

    # Products by brand
    por_marca = query("""
        SELECT marca, COUNT(*) as cantidad
        FROM productos WHERE activo=1
        GROUP BY marca ORDER BY cantidad DESC LIMIT 10
    """)

    # Price distribution
    rangos = query("""
        SELECT
            CASE
                WHEN precio_venta < 2000 THEN 'Menos de $2.000'
                WHEN precio_venta < 5000 THEN '$2.000 - $5.000'
                WHEN precio_venta < 10000 THEN '$5.000 - $10.000'
                WHEN precio_venta < 20000 THEN '$10.000 - $20.000'
                ELSE 'Más de $20.000'
            END as rango,
            COUNT(*) as cantidad
        FROM productos WHERE activo=1
        GROUP BY rango
    """)

    # Top 10 most expensive
    top10 = query("""
        SELECT nombre, precio_venta, categoria, marca
        FROM productos WHERE activo=1
        ORDER BY precio_venta DESC LIMIT 10
    """)

    # Recent price history
    historial = query("""
        SELECT h.fecha, p.nombre as producto_nombre, h.precio_anterior, h.precio_nuevo, h.motivo
        FROM historial_precios h JOIN productos p ON h.producto_id = p.id
        ORDER BY h.fecha DESC LIMIT 5
    """)

    return render_template("dashboard.html",
        total_productos=total_productos,
        total_clientes=total_clientes,
        total_pedidos=total_pedidos,
        valor_catalogo=valor_catalogo,
        por_categoria=por_categoria,
        por_marca=por_marca,
        rangos=rangos,
        top10=top10,
        historial=historial,
        now=datetime.now().isoformat()
    )

@app.route("/catalogo")
def catalogo():
    categoria = request.args.get("categoria", "")
    marca = request.args.get("marca", "")
    buscar = request.args.get("buscar", "")

    sql = "SELECT * FROM productos WHERE activo=1"
    args = []
    if categoria:
        sql += " AND categoria=?"; args.append(categoria)
    if marca:
        sql += " AND marca=?"; args.append(marca)
    if buscar:
        sql += " AND nombre LIKE ?"; args.append(f"%{buscar}%")
    sql += " ORDER BY categoria, nombre"

    productos = query(sql, args)
    categorias = query("SELECT DISTINCT categoria FROM productos WHERE activo=1 ORDER BY categoria")
    marcas = query("SELECT DISTINCT marca FROM productos WHERE activo=1 ORDER BY marca")

    return render_template("catalogo.html",
        productos=productos,
        categorias=categorias,
        marcas=marcas,
        filtro_cat=categoria,
        filtro_marca=marca,
        buscar=buscar
    )

@app.route("/catalogo/nuevo")
def catalogo_nuevo_redirect():
    return redirect(url_for("nuevo_producto"))

@app.route("/producto/nuevo", methods=["GET","POST"])
def nuevo_producto():
    if request.method == "POST":
        d = request.form
        precio_costo = float(d.get("precio_costo", 0) or 0)
        margen = float(d.get("margen", 30) or 30)
        precio_venta = precio_costo * (1 + margen/100) if precio_costo > 0 else float(d.get("precio_venta", 0) or 0)
        def pf(k): return float(d.get(k, 0) or 0)
        execute("""
            INSERT INTO productos (nombre, categoria, marca, codigo, precio_costo, margen, precio_venta,
                                   precio_x3, precio_x12, precio_x24, precio_x48, stock, updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (d["nombre"], d["categoria"], d["marca"], d.get("codigo",""),
              precio_costo, margen, precio_venta,
              pf("precio_x3"), pf("precio_x12"), pf("precio_x24"), pf("precio_x48"),
              int(d.get("stock",0) or 0), datetime.now().isoformat()))
        flash("Producto agregado correctamente", "success")
        return redirect(url_for("catalogo"))
    categorias = [r["categoria"] for r in query("SELECT DISTINCT categoria FROM productos ORDER BY categoria")]
    marcas = [r["marca"] for r in query("SELECT DISTINCT marca FROM productos ORDER BY marca")]
    return render_template("producto_form.html", producto=None, categorias=categorias, marcas=marcas)

@app.route("/producto/editar/<int:pid>", methods=["GET","POST"])
def editar_producto(pid):
    producto = query("SELECT * FROM productos WHERE id=?", [pid], one=True)
    if request.method == "POST":
        d = request.form
        precio_costo = float(d.get("precio_costo", 0) or 0)
        margen = float(d.get("margen", 30) or 30)
        precio_venta = precio_costo * (1 + margen/100) if precio_costo > 0 else float(d.get("precio_venta", 0) or 0)
        old_price = producto["precio_venta"]
        def pf(k): return float(d.get(k, 0) or 0)
        execute("""
            UPDATE productos SET nombre=?, categoria=?, marca=?, codigo=?,
                precio_costo=?, margen=?, precio_venta=?,
                precio_x3=?, precio_x12=?, precio_x24=?, precio_x48=?,
                stock=?, updated_at=?
            WHERE id=?
        """, (d["nombre"], d["categoria"], d["marca"], d.get("codigo",""),
              precio_costo, margen, precio_venta,
              pf("precio_x3"), pf("precio_x12"), pf("precio_x24"), pf("precio_x48"),
              int(d.get("stock",0) or 0), datetime.now().isoformat(), pid))
        if abs(old_price - precio_venta) > 0.01:
            execute("INSERT INTO historial_precios (producto_id, precio_anterior, precio_nuevo, motivo, fecha) VALUES (?,?,?,?,?)",
                    (pid, old_price, precio_venta, "Edición manual", datetime.now().isoformat()))
        flash("Producto actualizado", "success")
        return redirect(url_for("catalogo"))
    categorias = [r["categoria"] for r in query("SELECT DISTINCT categoria FROM productos ORDER BY categoria")]
    marcas = [r["marca"] for r in query("SELECT DISTINCT marca FROM productos ORDER BY marca")]
    return render_template("producto_form.html", producto=producto, categorias=categorias, marcas=marcas)

@app.route("/producto/eliminar/<int:pid>", methods=["POST"])
def eliminar_producto(pid):
    execute("UPDATE productos SET activo=0 WHERE id=?", [pid])
    flash("Producto eliminado", "warning")
    return redirect(url_for("catalogo"))

@app.route("/productos/importar", methods=["POST"])
def importar_productos():
    if "archivo" not in request.files:
        flash("No se seleccionó archivo", "danger")
        return redirect(url_for("catalogo"))
    f = request.files["archivo"]
    if not f.filename:
        flash("Archivo vacío", "danger")
        return redirect(url_for("catalogo"))

    modo = request.form.get("modo", "omitir")  # omitir | actualizar | duplicar

    try:
        filename = f.filename.lower()
        ahora = datetime.now().isoformat()

        # ── Leer filas ──────────────────────────────────────────────────────
        if filename.endswith(".csv"):
            import csv, io as _io
            content = f.read().decode("utf-8-sig")
            reader = csv.DictReader(_io.StringIO(content))
            rows = list(reader)
        else:
            import openpyxl
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c).strip().lower() if c is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append({headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)})

        # ── Mapeo flexible de columnas ──────────────────────────────────────
        COL_MAP = {
            "nombre":       ["nombre", "name", "producto", "descripcion", "descripción", "articulo", "artículo"],
            "categoria":    ["categoria", "categoría", "category", "rubro", "tipo"],
            "marca":        ["marca", "brand", "fabricante", "proveedor"],
            "codigo":       ["codigo", "código", "code", "sku", "ref", "referencia"],
            "precio_costo": ["precio_costo", "costo", "precio costo", "cost", "precio de costo"],
            "margen":       ["margen", "margin", "margen %", "margen%", "% margen"],
            "precio_venta": ["precio_venta", "precio", "price", "precio de venta", "pvp", "venta"],
            "stock":        ["stock", "cantidad", "qty", "inventario", "existencias"],
        }

        def find_col(row, field):
            for alias in COL_MAP[field]:
                for k in row.keys():
                    if k and k.strip().lower() == alias:
                        v = row[k]
                        return v if v is not None else ""
            return ""

        def to_float(v):
            try:
                return float(str(v).replace(",", ".").replace("$", "").strip())
            except:
                return 0.0

        def to_int(v):
            try:
                return int(float(str(v).replace(",", ".").strip()))
            except:
                return 0

        # ── Cargar nombres existentes para detección de duplicados ──────────
        conn = get_db()
        existentes = {r["nombre"].strip().lower(): r["id"]
                      for r in conn.execute("SELECT id, nombre FROM productos").fetchall()}

        agregados = 0
        actualizados = 0
        omitidos = 0
        vacias = 0

        batch_insert = []
        batch_update = []

        for row in rows:
            nombre = find_col(row, "nombre").strip()
            if not nombre:
                vacias += 1
                continue

            categoria   = find_col(row, "categoria") or "General"
            marca       = find_col(row, "marca") or ""
            codigo      = find_col(row, "codigo") or ""
            precio_costo= to_float(find_col(row, "precio_costo"))
            margen      = to_float(find_col(row, "margen")) or 30.0
            precio_venta= to_float(find_col(row, "precio_venta"))
            stock       = to_int(find_col(row, "stock"))

            # Calcular precio_venta si solo tiene costo y margen
            if precio_venta == 0 and precio_costo > 0:
                precio_venta = round(precio_costo * (1 + margen / 100), 2)
            # Calcular costo si solo tiene precio y margen
            if precio_costo == 0 and precio_venta > 0:
                precio_costo = round(precio_venta / (1 + margen / 100), 2)

            nombre_lower = nombre.lower()
            ya_existe = nombre_lower in existentes

            if ya_existe:
                if modo == "omitir":
                    omitidos += 1
                    continue
                elif modo == "actualizar":
                    batch_update.append((precio_costo, margen, precio_venta, stock, ahora, existentes[nombre_lower]))
                    actualizados += 1
                    continue
                # modo "duplicar" → cae al insert

            batch_insert.append((nombre, categoria, marca, codigo,
                                  precio_costo, margen, precio_venta,
                                  0, 0, 0, 0, stock, 1, ahora))
            agregados += 1

        # ── Insertar en batch ───────────────────────────────────────────────
        if batch_insert:
            conn.executemany("""
                INSERT INTO productos
                  (nombre, categoria, marca, codigo, precio_costo, margen, precio_venta,
                   precio_x3, precio_x12, precio_x24, precio_x48, stock, activo, updated_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, batch_insert)

        if batch_update:
            conn.executemany("""
                UPDATE productos SET precio_costo=?, margen=?, precio_venta=?, stock=?, updated_at=?
                WHERE id=?
            """, batch_update)

        conn.commit()
        conn.close()

        partes = [f"✓ {agregados} productos importados"]
        if actualizados: partes.append(f"{actualizados} actualizados")
        if omitidos:     partes.append(f"{omitidos} omitidos (ya existían)")
        if vacias:       partes.append(f"{vacias} filas vacías ignoradas")
        flash(" · ".join(partes), "success" if agregados + actualizados > 0 else "warning")

    except Exception as e:
        flash(f"Error procesando archivo: {e}", "danger")

    return redirect(url_for("catalogo"))

@app.route("/productos/plantilla")
def plantilla_productos():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Productos"

        headers = ["nombre", "categoria", "marca", "codigo",
                   "precio_costo", "margen", "precio_venta", "stock"]
        anchos  = [45, 20, 20, 12, 15, 10, 15, 8]

        for col, (h, w) in enumerate(zip(headers, anchos), 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", start_color="6B2D6B")
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[chr(64 + col)].width = w

        # Fila de ejemplo
        ejemplo = ["Sahumerio Saphirus Lavanda x12",
                   "Sahumerios", "Saphirus", "SAP-LAV-12",
                   3500, 30, 4550, 48]
        for col, val in enumerate(ejemplo, 1):
            ws.cell(row=2, column=col, value=val)

        # Nota en fila 4
        nota = ws.cell(row=4, column=1, value="⚠  Si ponés costo y margen, el precio_venta se calcula solo. Podés dejar precio_venta vacío.")
        nota.font = Font(italic=True, color="888888", size=9)
        ws.merge_cells("A4:H4")

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True,
                         download_name="plantilla_productos.xlsx")
    except Exception as e:
        flash(f"Error: {e}", "danger")
        return redirect(url_for("catalogo"))

# ─── PRECIOS ─────────────────────────────────────────────────────────────────

@app.route("/precios")
def precios():
    productos = query("SELECT * FROM productos WHERE activo=1 ORDER BY categoria, nombre")
    historial = query("""
        SELECT h.*, p.nombre as producto_nombre
        FROM historial_precios h JOIN productos p ON h.producto_id=p.id
        ORDER BY h.fecha DESC LIMIT 20
    """)
    return render_template("precios.html", productos=productos, historial=historial)

@app.route("/precios/actualizar", methods=["POST"])
def actualizar_precio():
    pid = request.form.get("producto_id")
    nuevo_pv = float(request.form.get("precio_venta", 0) or 0)
    motivo = request.form.get("motivo", "Ajuste manual")
    p = query("SELECT precio_venta FROM productos WHERE id=?", [pid], one=True)
    if p:
        execute("INSERT INTO historial_precios (producto_id, precio_anterior, precio_nuevo, motivo, fecha) VALUES (?,?,?,?,?)",
                (pid, p["precio_venta"], nuevo_pv, motivo, datetime.now().isoformat()))
        execute("UPDATE productos SET precio_venta=?, updated_at=? WHERE id=?",
                (nuevo_pv, datetime.now().isoformat(), pid))
        flash(f"Precio actualizado a ${nuevo_pv:,.0f}", "success")
    return redirect(url_for("precios"))

@app.route("/precios/lista", methods=["POST"])
def actualizar_lista():
    f = request.files.get("archivo")
    if not f or not f.filename:
        flash("No se seleccionó archivo", "danger")
        return redirect(url_for("precios"))
    try:
        import openpyxl
        wb = openpyxl.load_workbook(f, read_only=True)
        ws = wb.active
        headers = [str(c).strip().lower() if c else "" for c in next(ws.iter_rows(values_only=True))]
        def ci(alias):
            for a in alias:
                if a in headers: return headers.index(a)
            return None
        ci_nom = ci(["nombre","producto","name"])
        ci_pv  = ci(["precio_venta","precio","pvp","venta"])
        if ci_nom is None:
            flash("El archivo necesita una columna 'nombre'", "danger")
            return redirect(url_for("precios"))
        ahora = datetime.now().isoformat()
        actualizados = 0; no_encontrados = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[ci_nom]: continue
            nombre = str(row[ci_nom]).strip()
            nuevo_pv = None
            if ci_pv is not None:
                try: nuevo_pv = float(str(row[ci_pv]).replace("$","").replace(",",".").strip())
                except: pass
            if nuevo_pv is None: continue
            prod = query("SELECT id, precio_venta FROM productos WHERE nombre LIKE ?", [f"%{nombre}%"], one=True)
            if prod:
                old = prod["precio_venta"]
                execute("UPDATE productos SET precio_venta=?, updated_at=? WHERE id=?", (nuevo_pv, ahora, prod["id"]))
                if abs(old - nuevo_pv) > 0.01:
                    execute("INSERT INTO historial_precios (producto_id, precio_anterior, precio_nuevo, motivo, fecha) VALUES (?,?,?,?,?)",
                            (prod["id"], old, nuevo_pv, "Lista de proveedor", ahora))
                actualizados += 1
            else:
                no_encontrados.append(nombre)
        msg = f"✓ {actualizados} productos actualizados."
        if no_encontrados:
            msg += f" {len(no_encontrados)} no encontrados: {', '.join(no_encontrados[:5])}"
        flash(msg, "success" if actualizados > 0 else "warning")
    except Exception as e:
        flash(f"Error procesando archivo: {str(e)}", "danger")
    return redirect(url_for("precios"))

# ─── CLIENTES ────────────────────────────────────────────────────────────────

@app.route("/clientes")
def clientes():
    lista = query("SELECT * FROM clientes ORDER BY nombre")
    return render_template("clientes.html", clientes=lista)

@app.route("/cliente/nuevo", methods=["POST"])
def nuevo_cliente():
    d = request.form
    execute("INSERT INTO clientes (nombre, cuit, email, telefono, direccion, tipo, created_at) VALUES (?,?,?,?,?,?,?)",
            (d["nombre"], d.get("cuit",""), d.get("email",""), d.get("telefono",""), d.get("direccion",""), d.get("tipo","mayorista"), datetime.now().isoformat()))
    flash("Cliente agregado", "success")
    return redirect(url_for("clientes"))

@app.route("/cliente/eliminar/<int:cid>", methods=["POST"])
def eliminar_cliente(cid):
    execute("DELETE FROM clientes WHERE id=?", [cid])
    flash("Cliente eliminado", "warning")
    return redirect(url_for("clientes"))

@app.route("/clientes/importar", methods=["POST"])
def importar_clientes():
    if "archivo" not in request.files:
        flash("No se seleccionó archivo", "danger")
        return redirect(url_for("clientes"))
    f = request.files["archivo"]
    if not f.filename:
        flash("Archivo vacío", "danger")
        return redirect(url_for("clientes"))
    try:
        ahora = datetime.now().isoformat()
        agregados = 0; omitidos = 0; errores = []
        filename = f.filename.lower()
        if filename.endswith(".csv"):
            import csv, io as _io
            content = f.read().decode("utf-8-sig")
            reader = csv.DictReader(_io.StringIO(content))
            rows = list(reader)
        else:
            import openpyxl
            wb = openpyxl.load_workbook(f)
            ws = wb.active
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append({headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)})
        COL_MAP = {
            "nombre": ["nombre","name","razón social","razon social","cliente"],
            "cuit":   ["cuit","cuit/cuil","cuil"],
            "telefono": ["telefono","teléfono","tel","phone","celular"],
            "email":  ["email","correo","mail","e-mail"],
            "direccion": ["direccion","dirección","address","domicilio"],
            "tipo":   ["tipo","type","categoría","categoria"],
        }
        def find_col(row, field):
            for alias in COL_MAP[field]:
                for k, v in row.items():
                    if k and k.lower().strip() == alias: return v
            return ""
        TIPOS_VALIDOS = {"mayorista","minorista","revendedor"}
        nombres_existentes = {r["nombre"].lower() for r in query("SELECT nombre FROM clientes")}
        batch = []
        for row in rows:
            nombre = find_col(row, "nombre").strip()
            if not nombre: continue
            if nombre.lower() in nombres_existentes:
                omitidos += 1; continue
            tipo = find_col(row, "tipo").lower().strip()
            if tipo not in TIPOS_VALIDOS: tipo = "mayorista"
            batch.append((nombre, find_col(row,"cuit"), find_col(row,"email"),
                          find_col(row,"telefono"), find_col(row,"direccion"), tipo, ahora))
        if batch:
            conn = get_db()
            conn.executemany("INSERT INTO clientes (nombre,cuit,email,telefono,direccion,tipo,created_at) VALUES (?,?,?,?,?,?,?)", batch)
            conn.commit(); conn.close()
            agregados = len(batch)
        flash(f"✓ {agregados} clientes importados. {omitidos} omitidos (ya existían).", "success")
    except Exception as e:
        flash(f"Error: {str(e)}", "danger")
    return redirect(url_for("clientes"))

@app.route("/clientes/plantilla")
def plantilla_clientes():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = "Clientes"
        headers = ["nombre","cuit","telefono","email","direccion","tipo"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", start_color="6B2D6B")
            cell.alignment = Alignment(horizontal="center")
        ws.append(["Kiosco Don Pedro","20-12345678-9","1134567890","pedro@mail.com","Av. Siempre Viva 742","mayorista"])
        output = io.BytesIO()
        wb.save(output); output.seek(0)
        return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="plantilla_clientes.xlsx")
    except Exception as e:
        flash(f"Error: {e}", "danger")
        return redirect(url_for("clientes"))

# ─── PEDIDOS / DESCUENTOS ─────────────────────────────────────────────────────

@app.route("/pedidos")
def pedidos():
    lista = query("""
        SELECT p.*, c.nombre as cliente_nombre
        FROM pedidos p LEFT JOIN clientes c ON p.cliente_id=c.id
        ORDER BY p.id DESC
    """)
    clientes_list = query("SELECT * FROM clientes ORDER BY nombre")
    productos_list = query("SELECT * FROM productos WHERE activo=1 ORDER BY nombre")
    precio_reglas = query("SELECT * FROM precio_reglas ORDER BY cantidad_minima")
    return render_template("pedidos.html", pedidos=lista, clientes=clientes_list,
                           productos=productos_list, precio_reglas=precio_reglas)

@app.route("/config/precios", methods=["GET","POST"])
def config_precios():
    if request.method == "POST":
        data = request.get_json(force=True)
        conn = get_db()
        for item in data.get("reglas", []):
            rid = item.get("id"); val = item.get("descuento_pct")
            if rid is not None and val is not None:
                try:
                    pct = float(val)
                    conn.execute("UPDATE precio_reglas SET descuento_pct=? WHERE id=?", (pct, rid))
                except: pass
        conn.commit(); conn.close()
        return jsonify({"ok": True})
    reglas = query("SELECT * FROM precio_reglas ORDER BY cantidad_minima")
    return jsonify([dict(r) for r in reglas])

@app.route("/pedido/nuevo", methods=["POST"])
def nuevo_pedido():
    d = request.form
    cliente_id = d.get("cliente_id") or None
    pid = execute("INSERT INTO pedidos (cliente_id, fecha, estado, notas) VALUES (?,?,?,?)",
                  (cliente_id, datetime.now().isoformat(), "pendiente", d.get("notas","")))
    reglas = query("SELECT * FROM precio_reglas ORDER BY cantidad_minima DESC")
    productos_ids = request.form.getlist("producto_id[]")
    cantidades = request.form.getlist("cantidad[]")
    total = 0
    for prod_id, cant in zip(productos_ids, cantidades):
        if not prod_id: continue
        prod = query("SELECT precio_venta FROM productos WHERE id=?", [prod_id], one=True)
        if prod:
            qty = int(cant) if cant else 1
            precio = prod["precio_venta"]
            for r in reglas:
                if qty >= r["cantidad_minima"] and r["descuento_pct"] > 0:
                    precio = round(prod["precio_venta"] * (1 - r["descuento_pct"] / 100), 2)
                    break
            execute("INSERT INTO pedido_items (pedido_id, producto_id, cantidad, precio_unitario) VALUES (?,?,?,?)",
                    (pid, prod_id, qty, precio))
            total += precio * qty
    execute("UPDATE pedidos SET total=? WHERE id=?", (total, pid))
    flash(f"Pedido #{pid} creado por ${total:,.0f}", "success")
    return redirect(url_for("pedidos"))

@app.route("/pedido/estado/<int:pid>", methods=["POST"])
def cambiar_estado(pid):
    estado = request.form.get("estado")
    execute("UPDATE pedidos SET estado=? WHERE id=?", (estado, pid))
    return redirect(url_for("pedidos"))

@app.route("/pedido/<int:pid>/boleta")
def boleta(pid):
    pedido = query("SELECT p.*, c.nombre as cliente_nombre, c.cuit, c.direccion, c.telefono, c.tipo "
                   "FROM pedidos p LEFT JOIN clientes c ON p.cliente_id=c.id WHERE p.id=?", [pid], one=True)
    if not pedido:
        flash("Pedido no encontrado", "danger")
        return redirect(url_for("pedidos"))
    items = query("""
        SELECT pi.cantidad, pi.precio_unitario, p.nombre, p.marca, p.categoria
        FROM pedido_items pi JOIN productos p ON pi.producto_id=p.id
        WHERE pi.pedido_id=?
    """, [pid])
    subtotal = sum((i["precio_unitario"] or 0) * (i["cantidad"] or 1) for i in items)
    iva = round(subtotal * 0.21, 2)
    total = subtotal + iva
    return render_template("boleta.html", pedido=pedido, items=items,
                           subtotal=subtotal, iva=iva, total=total,
                           numero=f"{pid:06d}", fecha=datetime.now().strftime("%d/%m/%Y %H:%M"))

# ─── API ─────────────────────────────────────────────────────────────────────

@app.route("/api/productos")
def api_productos():
    productos = query("SELECT * FROM productos WHERE activo=1 ORDER BY nombre")
    return jsonify([dict(p) for p in productos])

@app.route("/api/incrementar_costos", methods=["POST"])
def incrementar_costos():
    d = request.json
    pct = float(d.get("porcentaje", 0))
    if pct <= 0: return jsonify({"ok": False, "error": "Porcentaje inválido"})
    productos = query("SELECT * FROM productos WHERE activo=1 AND precio_costo > 0")
    ahora = datetime.now().isoformat()
    for p in productos:
        nuevo_costo = p["precio_costo"] * (1 + pct / 100)
        margen = p["margen"] or 30
        nuevo_pv = nuevo_costo * (1 + margen / 100)
        execute("UPDATE productos SET precio_costo=?, precio_venta=?, updated_at=? WHERE id=?",
                (nuevo_costo, nuevo_pv, ahora, p["id"]))
        execute("INSERT INTO historial_precios (producto_id, precio_anterior, precio_nuevo, motivo, fecha) VALUES (?,?,?,?,?)",
                (p["id"], p["precio_venta"], nuevo_pv, f"Incremento global +{pct}%", ahora))
    return jsonify({"ok": True, "actualizados": len(productos)})

@app.route("/api/dashboard_data")
def api_dashboard():
    por_cat = query("SELECT categoria, COUNT(*) as n FROM productos WHERE activo=1 GROUP BY categoria ORDER BY n DESC")
    return jsonify({"por_categoria": [dict(r) for r in por_cat]})

@app.route("/exportar/excel")
def exportar_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        productos = query("SELECT * FROM productos WHERE activo=1 ORDER BY categoria, nombre")
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Catálogo"
        headers = ["ID","Nombre","Categoría","Marca","Costo","Margen %","Precio Venta","Stock"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", start_color="1B3A4B")
            cell.alignment = Alignment(horizontal="center")
        for i, p in enumerate(productos, 2):
            ws.cell(row=i, column=1, value=p["id"])
            ws.cell(row=i, column=2, value=p["nombre"])
            ws.cell(row=i, column=3, value=p["categoria"])
            ws.cell(row=i, column=4, value=p["marca"])
            ws.cell(row=i, column=5, value=p["precio_costo"]).number_format = "$ #,##0"
            ws.cell(row=i, column=6, value=float(p["margen"] or 0) / 100).number_format = "0.0%"
            ws.cell(row=i, column=7, value=p["precio_venta"]).number_format = "$ #,##0"
            ws.cell(row=i, column=8, value=p["stock"])
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 20
        output = io.BytesIO(); wb.save(output); output.seek(0)
        return send_file(output,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="catalogo_oxuma.xlsx")
    except Exception as e:
        flash(f"Error exportando: {e}", "danger")
        return redirect(url_for("catalogo"))

# ─── COMBOS / PROMOCIONES ────────────────────────────────────────────────────

def _get_combo_full(combo_id):
    combo = query("SELECT * FROM combos WHERE id=?", [combo_id], one=True)
    if not combo: return None
    items = query("""
        SELECT ci.cantidad, p.nombre, p.precio_venta, p.marca, p.categoria, p.id as pid
        FROM combo_items ci JOIN productos p ON ci.producto_id=p.id
        WHERE ci.combo_id=?
    """, [combo_id])
    return {"combo": dict(combo), "items": [dict(i) for i in items]}

@app.route("/combos")
def combos():
    lista = query("SELECT * FROM combos WHERE activo=1 ORDER BY created_at DESC")
    combos_full = []
    for c in lista:
        items = query("""
            SELECT ci.cantidad, p.nombre, p.precio_venta, p.marca
            FROM combo_items ci JOIN productos p ON ci.producto_id=p.id
            WHERE ci.combo_id=?
        """, [c["id"]])
        precio_orig = sum((i["precio_venta"] or 0) * (i["cantidad"] or 1) for i in items)
        combos_full.append({"combo": c, "items": items, "precio_original": precio_orig,
                            "ahorro": precio_orig - (c["precio_combo"] or 0)})
    productos = query("SELECT id, nombre, precio_venta, categoria, marca FROM productos WHERE activo=1 ORDER BY nombre")
    return render_template("combos.html", combos=combos_full, productos=productos)

@app.route("/combo/nuevo", methods=["POST"])
def nuevo_combo():
    d = request.form
    precio = float(d.get("precio_combo", 0) or 0)
    cid = execute("INSERT INTO combos (nombre, descripcion, precio_combo, activo, ia_generado, created_at) VALUES (?,?,?,1,0,?)",
                  (d["nombre"], d.get("descripcion",""), precio, datetime.now().isoformat()))
    pids = request.form.getlist("producto_id[]")
    cnts = request.form.getlist("cantidad[]")
    for pid, cnt in zip(pids, cnts):
        if pid:
            execute("INSERT INTO combo_items (combo_id, producto_id, cantidad) VALUES (?,?,?)",
                    (cid, pid, int(cnt) if cnt else 1))
    flash("Combo creado", "success")
    return redirect(url_for("combos"))

@app.route("/combo/eliminar/<int:cid>", methods=["POST"])
def eliminar_combo(cid):
    execute("UPDATE combos SET activo=0 WHERE id=?", [cid])
    execute("DELETE FROM combo_items WHERE combo_id=?", [cid])
    flash("Combo eliminado", "warning")
    return redirect(url_for("combos"))

@app.route("/combos/generar-ia", methods=["POST"])
def generar_combos_ia():
    import random
    old_ia = query("SELECT id FROM combos WHERE ia_generado=1 AND activo=1")
    for c in old_ia:
        execute("UPDATE combos SET activo=0 WHERE id=?", [c["id"]])
        execute("DELETE FROM combo_items WHERE combo_id=?", [c["id"]])
    productos = query("SELECT * FROM productos WHERE activo=1 AND precio_venta > 0")
    if not productos:
        flash("No hay productos en el catálogo para generar combos", "warning")
        return redirect(url_for("combos"))
    por_cat = {}
    for p in productos:
        cat = p["categoria"] or "General"
        por_cat.setdefault(cat, []).append(p)
    categorias = list(por_cat.keys())
    ahora = datetime.now().isoformat()
    generados = 0
    DESCUENTO_2 = 0.10; DESCUENTO_3 = 0.15; DESCUENTO_4 = 0.20
    NOMBRES = {
        2: ["Kit Dúo {cat}","Pack {cat} Esencial","Combo {cat} Básico"],
        3: ["Kit Completo {cat}","Pack Experiencia {cat}","Combo {cat} Premium"],
        4: ["Mega Kit {cat}","Pack Profesional {cat}","Colección {cat}"],
    }
    DESCRIPCIONES = {
        2: "Combo ideal para comenzar. Ahorrás comprando juntos.",
        3: "Set completo para disfrutar en casa. Mejor precio por combo.",
        4: "Kit profesional con todo lo necesario. Máximo ahorro garantizado.",
    }
    for cat in categorias:
        prods = por_cat[cat]
        if len(prods) < 2: continue
        for size in [2, 3, 4]:
            if len(prods) < size: continue
            desc = DESCUENTO_2 if size == 2 else (DESCUENTO_3 if size == 3 else DESCUENTO_4)
            muestra = random.sample(prods, size)
            precio_orig = sum(p["precio_venta"] for p in muestra)
            precio_combo = round(precio_orig * (1 - desc))
            nombre = random.choice(NOMBRES[size]).format(cat=cat)
            cid = execute("INSERT INTO combos (nombre, descripcion, precio_combo, activo, ia_generado, created_at) VALUES (?,?,?,1,1,?)",
                         (nombre, DESCRIPCIONES[size], precio_combo, ahora))
            for p in muestra:
                execute("INSERT INTO combo_items (combo_id, producto_id, cantidad) VALUES (?,?,1)", (cid, p["id"]))
            generados += 1
    prods_premium = sorted(productos, key=lambda x: x["precio_venta"], reverse=True)[:20]
    if len(prods_premium) >= 2:
        regalo = random.sample(prods_premium, 2)
        precio_orig = sum(p["precio_venta"] for p in regalo)
        precio_combo = round(precio_orig * (1 - DESCUENTO_2))
        cid = execute("INSERT INTO combos (nombre, descripcion, precio_combo, activo, ia_generado, created_at) VALUES (?,?,?,1,1,?)",
                     ("Pack Regalo Premium", "Combo ideal para regalar. Productos seleccionados del catálogo.", precio_combo, ahora))
        for p in regalo:
            execute("INSERT INTO combo_items (combo_id, producto_id, cantidad) VALUES (?,?,1)", (cid, p["id"]))
        generados += 1
    flash(f"✨ Se generaron {generados} combos nuevos con IA", "success")
    return redirect(url_for("combos"))

# ─── ASISTENTE IA ────────────────────────────────────────────────────────────

@app.route("/asistente", methods=["POST"])
def asistente():
    data = request.get_json(force=True)
    pregunta = (data.get("pregunta") or "").lower().strip()

    def pesos(n): return f"${n:,.0f}"

    respuesta = ""

    if any(w in pregunta for w in ["cuántos productos","cuantos productos","total de productos","cantidad de productos"]):
        n = query("SELECT COUNT(*) as n FROM productos WHERE activo=1", one=True)["n"]
        respuesta = f"Tenés {n} productos activos en el catálogo."

    elif any(w in pregunta for w in ["categoría","categoria","categorías","categorias"]) and any(w in pregunta for w in ["cuántas","cuantas","lista","cuales","qué","que"]):
        cats = query("SELECT categoria, COUNT(*) as n FROM productos WHERE activo=1 GROUP BY categoria ORDER BY n DESC")
        lines = [f"• {c['categoria']}: {c['n']} productos" for c in cats]
        respuesta = f"Hay {len(cats)} categorías:\n" + "\n".join(lines)

    elif any(w in pregunta for w in ["más caro","mas caro","mayor precio","precio más alto"]):
        r = query("SELECT nombre, precio_venta, marca FROM productos WHERE activo=1 ORDER BY precio_venta DESC LIMIT 5")
        lines = [f"{i+1}. {p['nombre']} — {pesos(p['precio_venta'])}" for i, p in enumerate(r)]
        respuesta = "Los 5 más caros:\n" + "\n".join(lines)

    elif any(w in pregunta for w in ["más barato","mas barato","precio más bajo","precio mas bajo"]):
        r = query("SELECT nombre, precio_venta FROM productos WHERE activo=1 AND precio_venta > 0 ORDER BY precio_venta ASC LIMIT 5")
        lines = [f"{i+1}. {p['nombre']} — {pesos(p['precio_venta'])}" for i, p in enumerate(r)]
        respuesta = "Los 5 más económicos:\n" + "\n".join(lines)

    elif any(w in pregunta for w in ["precio promedio","promedio de precio"]):
        r = query("SELECT AVG(precio_venta) as avg FROM productos WHERE activo=1 AND precio_venta > 0", one=True)
        respuesta = f"Precio promedio de venta: {pesos(r['avg'])}."

    elif any(w in pregunta for w in ["valor del catálogo","valor catalogo","valor total"]):
        r = query("SELECT SUM(precio_venta) as s FROM productos WHERE activo=1", one=True)
        respuesta = f"Valor total del catálogo: {pesos(r['s'] or 0)}."

    elif "stock" in pregunta and any(w in pregunta for w in ["bajo","poco","sin","agotado"]):
        r = query("SELECT nombre, stock FROM productos WHERE activo=1 AND stock < 5 ORDER BY stock ASC LIMIT 10")
        if r:
            lines = [f"• {p['nombre']}: {p['stock']} u." for p in r]
            respuesta = f"⚠️ {len(r)} productos con stock bajo:\n" + "\n".join(lines)
        else:
            respuesta = "✅ Todos los productos tienen stock suficiente."

    elif any(w in pregunta for w in ["buscar","busca","busco"]):
        for stop in ["buscar","busca","busco","producto","el","la","los","las","un","una"]:
            pregunta = pregunta.replace(stop,"").strip()
        if pregunta:
            r = query("SELECT nombre, precio_venta FROM productos WHERE activo=1 AND nombre LIKE ? ORDER BY nombre LIMIT 8", [f"%{pregunta}%"])
            if r:
                lines = [f"• {p['nombre']} — {pesos(p['precio_venta'])}" for p in r]
                respuesta = f"Encontré {len(r)} productos:\n" + "\n".join(lines)
            else:
                respuesta = f"No encontré productos que coincidan con '{pregunta}'."
        else:
            respuesta = "¿Qué producto querés buscar?"

    elif any(w in pregunta for w in ["cuántos pedidos","cuantos pedidos","pedidos hoy"]):
        hoy = datetime.now().strftime("%Y-%m-%d")
        r = query("SELECT COUNT(*) as n, SUM(total) as s FROM pedidos WHERE fecha LIKE ?", [f"{hoy}%"], one=True)
        n = r["n"] or 0; s = r["s"] or 0
        respuesta = f"Hoy hay {n} pedido{'s' if n!=1 else ''}, total: {pesos(s)}."

    elif any(w in pregunta for w in ["ventas del mes","ventas este mes","cuánto vendí","cuanto vendi"]):
        mes = datetime.now().strftime("%Y-%m")
        r = query("SELECT COUNT(*) as n, SUM(total) as s FROM pedidos WHERE fecha LIKE ?", [f"{mes}%"], one=True)
        n = r["n"] or 0; s = r["s"] or 0
        respuesta = f"Este mes: {n} pedido{'s' if n!=1 else ''} por {pesos(s)}."

    elif any(w in pregunta for w in ["pendiente","pendientes","sin entregar"]):
        r = query("SELECT COUNT(*) as n FROM pedidos WHERE estado='pendiente'", one=True)
        respuesta = f"Hay {r['n']} pedido{'s' if r['n']!=1 else ''} pendiente{'s' if r['n']!=1 else ''}."

    elif any(w in pregunta for w in ["último pedido","ultimo pedido"]):
        r = query("""SELECT p.id, p.total, p.estado, c.nombre as cliente
                     FROM pedidos p LEFT JOIN clientes c ON p.cliente_id=c.id
                     ORDER BY p.id DESC LIMIT 1""", one=True)
        if r: respuesta = f"Último pedido #{r['id']} para {r['cliente'] or 'sin nombre'}, {pesos(r['total'])}, {r['estado']}."
        else: respuesta = "No hay pedidos aún."

    elif any(w in pregunta for w in ["cuántos clientes","cuantos clientes"]):
        n = query("SELECT COUNT(*) as n FROM clientes", one=True)["n"]
        respuesta = f"Tenés {n} clientes registrados."

    elif any(w in pregunta for w in ["descuento","descuentos","precio por cantidad"]):
        reglas = query("SELECT * FROM precio_reglas ORDER BY cantidad_minima")
        lines = [f"• x{r['cantidad_minima']} o más: {r['descuento_pct']:.0f}% de descuento" for r in reglas]
        respuesta = "Descuentos por volumen:\n" + "\n".join(lines)

    elif any(w in pregunta for w in ["marca","marcas"]):
        r = query("SELECT marca, COUNT(*) as n FROM productos WHERE activo=1 GROUP BY marca ORDER BY n DESC LIMIT 8")
        lines = [f"• {m['marca']}: {m['n']} productos" for m in r]
        respuesta = f"Tenés {len(r)} marcas:\n" + "\n".join(lines)

    elif any(w in pregunta for w in ["resumen","estado","cómo estamos","como estamos"]):
        prods = query("SELECT COUNT(*) as n FROM productos WHERE activo=1", one=True)["n"]
        clientes_n = query("SELECT COUNT(*) as n FROM clientes", one=True)["n"]
        pend = query("SELECT COUNT(*) as n FROM pedidos WHERE estado='pendiente'", one=True)["n"]
        mes = datetime.now().strftime("%Y-%m")
        ventas = query("SELECT SUM(total) as s FROM pedidos WHERE fecha LIKE ?", [f"{mes}%"], one=True)["s"] or 0
        respuesta = (f"📊 Resumen Oxuma:\n• Productos: {prods}\n• Clientes: {clientes_n}\n"
                     f"• Pedidos pendientes: {pend}\n• Ventas del mes: {pesos(ventas)}")

    elif any(w in pregunta for w in ["ayuda","help","qué podés","que podes"]):
        respuesta = ("Puedo responder:\n• ¿Cuántos productos hay?\n• Producto más caro/barato\n"
                     "• ¿Qué categorías tengo?\n• Ventas del mes / pedidos pendientes\n"
                     "• ¿Cuántos clientes tengo?\n• Descuentos por volumen\n• Buscar [producto]\n• Resumen general")
    else:
        respuesta = "No entendí la pregunta 🤔 Decí 'ayuda' para ver qué puedo responder."

    return jsonify({"respuesta": respuesta})


@app.route("/api/stats")
def api_stats():
    prods = query("SELECT COUNT(*) as n FROM productos WHERE activo=1", one=True)["n"]
    clientes_n = query("SELECT COUNT(*) as n FROM clientes", one=True)["n"]
    pend = query("SELECT COUNT(*) as n FROM pedidos WHERE estado='pendiente'", one=True)["n"]
    hoy = datetime.now().strftime("%Y-%m-%d")
    hoy_total = query("SELECT SUM(total) as s FROM pedidos WHERE fecha LIKE ?", [f"{hoy}%"], one=True)["s"] or 0
    return jsonify({"productos": prods, "clientes": clientes_n, "pendientes": pend, "ventas_hoy": hoy_total})


# ─── MAIN ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    init_db()
    migrate_db()
    app.run(host="0.0.0.0", debug=False, port=5000)
